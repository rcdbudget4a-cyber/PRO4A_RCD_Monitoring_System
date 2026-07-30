from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

UPLOAD = Path("/workspace/scratch/7517778b1e00/upload")
OUTPUT = Path("/workspace/sites/pro4a-rcd-system/app/retirees-data.ts")
RANKS = ["PBGEN", "PLTCOL", "PMAJ", "PCAPT", "PCPT", "PLT", "PEMS", "PSMS",
         "PCMS", "PMSg", "PSSg", "PCpl", "Pat", "NUP"]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%B %d, %Y")
    return re.sub(r"\s+", " ", str(value)).strip()


def split_rank(full_name: str) -> tuple[str, str]:
    for rank in RANKS:
        if full_name.upper().startswith(rank.upper() + " "):
            return rank, full_name[len(rank):].strip()
    parts = full_name.split(maxsplit=1)
    return parts[0], parts[1] if len(parts) > 1 else ""


def normalized_date(value) -> tuple[str, str]:
    if isinstance(value, datetime):
        return value.date().isoformat(), value.strftime("%B %d, %Y")
    text = clean(value)
    normalized = re.sub(r"\s+", " ", text.replace("-", " ")).strip()
    for fmt in ["%b %d %y", "%B %d %y", "%b %d %Y", "%B %d %Y"]:
        try:
            parsed = datetime.strptime(normalized, fmt)
            return parsed.date().isoformat(), parsed.strftime("%B %d, %Y")
        except ValueError:
            pass
    return text, text


def classify(cal: str, lump_sum: str, remarks: str) -> str:
    joined = f"{cal} {lump_sum} {remarks}".lower()
    if "under bos" in joined:
        return "Under BOS"
    if ("complete" in cal.lower() or "completed" in cal.lower()) and \
            "waiting for the release of clearance" not in remarks.lower():
        return "Complete"
    if "waiting for the release of clearance" in remarks.lower():
        return "Pending Clearance"
    if cal or lump_sum:
        return "Lacking Requirements"
    return "For Endorsement"


records = []
sources = [
    ("Compulsory Retirees-2025.xlsx", "Compulsory 2025", 2025, 4, 40, "CY 2025"),
    ("Compulsory Retirees-2026.xlsx", "Compulsory 2026", 2026, 4, 42, "As of July 15, 2026"),
]

for filename, sheet_name, year, first_row, last_row, coverage in sources:
    ws = load_workbook(UPLOAD / filename, data_only=True, read_only=True)[sheet_name]
    sequence = 0
    for row_number in range(first_row, last_row + 1):
        row = [ws.cell(row_number, col).value for col in range(1, 10)]
        if not clean(row[0]).isdigit():
            continue
        sequence += 1
        if year == 2025:
            full_name, retirement_value = clean(row[2]), row[4]
            cal, lump_sum, remarks = clean(row[5]), clean(row[6]), clean(row[7])
        else:
            full_name, retirement_value = clean(row[1]), row[2]
            cal, lump_sum, remarks = clean(row[3]), clean(row[4]), clean(row[5])
        rank, name = split_rank(full_name)
        retirement_date, retirement_display = normalized_date(retirement_value)
        records.append({
            "id": f"RET-{year}-{sequence:03d}", "year": year, "rank": rank, "name": name,
            "retirementDate": retirement_date, "retirementDisplay": retirement_display,
            "calRequirements": cal or "Not indicated",
            "lumpSumRequirements": lump_sum or "Not indicated",
            "status": classify(cal, lump_sum, remarks),
            "remarks": remarks or "No remarks indicated", "sourceCoverage": coverage,
        })

OUTPUT.write_text(
    "// Generated from the CY 2025 and CY 2026 compulsory retiree workbooks.\n"
    "// Contact numbers are intentionally excluded from the public website.\n"
    f"export const retireeRecords = {json.dumps(records, ensure_ascii=False, indent=2)};\n",
    encoding="utf-8",
)
print(json.dumps({
    "total": len(records),
    "byYear": {str(year): sum(r["year"] == year for r in records) for year in [2025, 2026]},
    "byStatus": {status: sum(r["status"] == status for r in records)
                 for status in sorted({r["status"] for r in records})},
}, indent=2))
