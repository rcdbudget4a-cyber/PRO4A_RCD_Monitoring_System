from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

UPLOAD = Path("/workspace/scratch/7517778b1e00/upload")
OUTPUT = Path("/workspace/sites/pro4a-rcd-system/app/claims-data.ts")

RANKS = [
    "PMAJ", "PLT", "PEMS", "PCMS", "PMSg", "PSMS", "PSSg", "PCpl",
    "Patrolman", "Pat", "PO1", "PO2", "PO3", "SPO1", "SPO2", "SPO3",
]


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def is_sequence(value) -> bool:
    return isinstance(value, (int, float)) or clean(value).isdigit()


def split_rank(full_name: str) -> tuple[str, str]:
    for rank in RANKS:
        if full_name.upper().startswith(rank.upper() + " "):
            return rank, full_name[len(rank):].strip()
    parts = full_name.split(maxsplit=1)
    return (parts[0], parts[1] if len(parts) > 1 else "")


def unit_group(unit: str) -> str:
    for name in ["Cavite PPO", "Laguna PPO", "Batangas PPO", "Rizal PPO", "Quezon PPO", "RMFB 4A"]:
        if name.lower().replace(" ", "") in unit.lower().replace(" ", ""):
            return name
    if "RPDEU" in unit.upper() or "RHQ" in unit.upper():
        return "RHQ"
    return "PRO 4A"


def iso_date(value: str) -> str:
    source = value.split("(")[0].strip().replace("July 8 2026", "July 8, 2026")
    for fmt in ["%B %d, %Y", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.strptime(source, fmt).date().isoformat()
        except ValueError:
            pass
    return source


def overall_status(text: str) -> tuple[str, str]:
    low = text.lower()
    if "on hold" in low or "for completion" in low:
        return "Document Completion", "Pending"
    if "received php" in low:
        return "Benefits Released", "Completed"
    if "approved" in low or "to receive" in low:
        return "Benefits Processing", "In Process"
    return "Document Review", "For Review"


def first_sheet(filename: str):
    return load_workbook(UPLOAD / filename, data_only=True, read_only=True).worksheets[0]


records: list[dict] = []

# KIPO: primary rows begin at 7 and end before the recapitulation.
for filename, year in [("KIPO 2025.xlsx", 2025), ("KIPO2026.xlsx", 2026)]:
    ws = first_sheet(filename)
    sequence = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not is_sequence(row[0]):
            continue
        sequence += 1
        full_name, unit, incident = clean(row[1]), clean(row[2]), clean(row[3])
        rank, name = split_rank(full_name)
        benefits = {
            "pnpSfa": clean(row[4]), "cal": clean(row[5]), "promotion": clean(row[6]),
            "awards": clean(row[7]), "napolcom": clean(row[8]), "psmbfi": clean(row[12]),
            "psfSfa": clean(row[13]), "education": clean(row[14]),
            "philHealth": clean(row[15]), "others": clean(row[16]),
        }
        status_text = " ".join(benefits.values())
        stage, status = overall_status(status_text)
        records.append({
            "id": f"KIPO-{year}-{sequence:03d}", "type": "KIPO", "year": year,
            "rank": rank, "name": name, "province": unit_group(unit), "office": unit,
            "stage": stage, "status": status, "date": iso_date(incident),
            "dateDisplay": incident, "sourceCoverage": "As of July 6, 2026",
            "injury": "", "benefits": benefits,
        })

# WIPO: 44 CY 2025 primary rows and 13 CY 2026 primary rows.
for filename, year, start, end in [
    ("WIPO2025.xlsx", 2025, 7, 50),
    ("WIPO2026.xlsx", 2026, 7, 19),
]:
    ws = first_sheet(filename)
    previous_date = ""
    previous_unit = ""
    sequence = 0
    for row_number in range(start, end + 1):
        row = [ws.cell(row_number, col).value for col in range(1, 14)]
        if not is_sequence(row[0]):
            continue
        sequence += 1
        full_name = clean(row[1])
        incident = clean(row[2]) or previous_date
        unit = clean(row[3]) or previous_unit
        previous_date, previous_unit = incident, unit
        rank, name = split_rank(full_name)
        benefits = {
            "rhe": clean(row[5]), "specialPromotion": clean(row[6]),
            "awards": clean(row[7]), "scholarship": clean(row[8]),
            "psmbfi": clean(row[9]), "psfSfa": clean(row[10]),
            "others": clean(row[11]),
        }
        stage, status = overall_status(benefits["psfSfa"])
        records.append({
            "id": f"WIPO-{year}-{sequence:03d}", "type": "WIPO", "year": year,
            "rank": rank, "name": name, "province": unit_group(unit), "office": unit,
            "stage": stage, "status": status, "date": iso_date(incident),
            "dateDisplay": incident, "sourceCoverage": "As of July 6, 2026",
            "injury": clean(row[12]), "benefits": benefits,
        })

payload = json.dumps(records, ensure_ascii=False, indent=2)
OUTPUT.write_text(
    "// Generated from the four user-provided final KIPO/WIPO workbooks.\n"
    "// Source coverage: As of July 6, 2026.\n"
    f"export const seedClaims = {payload};\n",
    encoding="utf-8",
)
print(json.dumps({
    "total": len(records),
    "KIPO": sum(r["type"] == "KIPO" for r in records),
    "WIPO": sum(r["type"] == "WIPO" for r in records),
    "byYear": {
        str(y): sum(r["year"] == y for r in records)
        for y in sorted({r["year"] for r in records})
    },
    "output": str(OUTPUT),
}, indent=2))
